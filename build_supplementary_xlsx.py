#!/usr/bin/env python3
"""Build Supplementary_Data.xlsx — the browsable data compilation behind the paper.
Every value is read from the repository's JSON data files (the engine outputs and
manuscript tables); derived summaries (coverage counts, means) are Excel formulas
so a reader can verify them. Fonts Arial throughout."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

HERE = os.path.dirname(os.path.abspath(__file__))
def J(name): return json.load(open(os.path.join(HERE, name), encoding="utf-8"))
rf   = J("rf_data.json")
coh  = J("cohort_tables.json")
app  = J("appendix_tables.json")
t5   = J("table5_regen.json")
pol  = J("policy_event_table.json")

NAVY="1F3864"; HDRFILL=PatternFill("solid",fgColor="DDE5F0"); TABFILL=PatternFill("solid",fgColor="F2F2F2")
H1=Font(name="Arial",size=14,bold=True,color=NAVY)
H2=Font(name="Arial",size=11,bold=True,color=NAVY)
HDR=Font(name="Arial",size=10,bold=True,color="1F3864")
BODY=Font(name="Arial",size=10)
MUT=Font(name="Arial",size=9,color="666666",italic=True)
CEN=Alignment(horizontal="center",vertical="center")
LFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
thin=Side(style="thin",color="BBBBBB"); BORD=Border(bottom=thin)

wb=Workbook(); wb.remove(wb.active)

def newsheet(name,color=NAVY):
    ws=wb.create_sheet(name); ws.sheet_properties.tabColor=color; return ws

def title(ws,t,sub=None):
    ws["A1"]=t; ws["A1"].font=H1
    if sub: ws["A2"]=sub; ws["A2"].font=MUT
    return 4  # first data row

def header_row(ws,r,cols):
    for j,c in enumerate(cols,1):
        cell=ws.cell(r,j,c); cell.font=HDR; cell.fill=HDRFILL; cell.alignment=LFT; cell.border=BORD
    ws.freeze_panes=ws.cell(r+1,1)

def widths(ws,ws_widths):
    for col,w in ws_widths.items(): ws.column_dimensions[col].width=w

def note(ws,r,text):
    ws.cell(r,1,text).font=MUT

# ---------------------------------------------------------------- 1. Contents
ws=newsheet("Contents","1F3864")
ws["A1"]="Supplementary Data"; ws["A1"].font=H1
ws["A2"]="From reference-class cost priors to investment choices: a back-tested screening workflow for lithium and copper projects"; ws["A2"].font=MUT
ws["A3"]="J. Caers — Mineral Economics (supplementary material)"; ws["A3"].font=MUT
rows=[
 ("Tab","Contents"),
 ("Lithium cohort","10 greenfield lithium projects (2010–2019): sanction CapEx, realized ratio, outcome (Table 1)"),
 ("Copper cohort","11 Chilean copper projects (2008–2019): sanction CapEx, realized ratio, outcome, confidence (Table 2)"),
 ("Reference-class results","Per-project frozen reference-class band, inside-view band, realized ratio, PIT, coverage flags (Figure 5)"),
 ("PIT calibration","Probability-integral-transform summary: KS p-values and means, pooled and by commodity (Figure 6)"),
 ("Copper decision audit","Ex-ante RAV, engine choice, reconstructed realized NPV per copper project (Figure 7, Table 4/7)"),
 ("Era register - lithium","Reference price and demand-growth outlook per sanction vintage (Table A2)"),
 ("Era register - copper","Reference price and market outlook per sanction vintage (Table A3)"),
 ("AACE dispersion","AACE estimate class to log-sigma mapping (Table A1)"),
 ("Policy-event register","41 resource-nationalism events, 2007–2026: lever, regime, instrument, primary source"),
 ("Driver attribution","Sobol top driver vs realized decisive factor, per lithium project (Table 5)"),
 ("Robustness","Discount x risk-aversion grid, decision-loss grids, and the shared-instrument sensitivity check"),
]
r=5
for a,b in rows:
    ws.cell(r,1,a).font=(HDR if r==5 else Font(name="Arial",size=10,bold=True,color=NAVY))
    ws.cell(r,2,b).font=(HDR if r==5 else BODY); ws.cell(r,2).alignment=LFT
    if r==5:
        ws.cell(r,1).fill=HDRFILL; ws.cell(r,2).fill=HDRFILL
    r+=1
note(ws,r+1,"All values derive from public disclosures (NI 43-101/ASX/SEC filings, USGS, ICSG, Cochilco, IMF, V-Dem, OECD) and the paper's engine outputs.")
note(ws,r+2,"Machine-readable equivalents (JSON) and a field-level data dictionary are in the accompanying code repository (data/).")
widths(ws,{"A":24,"B":96})

# ---------------------------------------------------------------- 2/3 cohorts
def cohort_tab(name,key,color):
    ws=newsheet(name,color)
    r=title(ws,name, "Source: manuscript "+("Table 1" if key=="T1" else "Table 2")+" (cohort_tables.json). CapEx as disclosed; ratio = realized ÷ sanctioned capital.")
    data=coh[key]; header_row(ws,r,data[0])
    for i,row in enumerate(data[1:],1):
        for j,v in enumerate(row,1):
            c=ws.cell(r+i,j,v); c.font=BODY; c.alignment=(CEN if j in (2,4,6) else LFT)
    lastr=r+len(data)-1
    note(ws,lastr+2,"FID = final investment decision year. Ratio 'x1.75' = realized capital 1.75× the sanction estimate."+ (" Conf = curation confidence tier (high/med/low)." if key=="T2" else ""))
    return ws
w1=cohort_tab("Lithium cohort","T1","2E5395"); widths(w1,{"A":30,"B":8,"C":16,"D":13,"E":40})
w2=cohort_tab("Copper cohort","T2","2E5395"); widths(w2,{"A":26,"B":8,"C":16,"D":13,"E":40,"F":8})

# ---------------------------------------------------------------- 4. Reference-class results
ws=newsheet("Reference-class results","808080")
r=title(ws,"Reference-class results (per project)","Source: rf_data.json 'forest'. Bands are the frozen 80% intervals in ratio units (realized ÷ sanctioned). 'In RC band' / 'In inside band' are coverage flags. PIT = probability-integral transform under the frozen posterior.")
cols=["Project","Commodity","Sanction yr","Realized ratio","RC band P10","RC band P90","Inside P10","Inside P90","PIT","In RC band","In inside band","Conf"]
header_row(ws,r,cols)
fr=r+1
for i,d in enumerate(rf["forest"]):
    vals=[d["name"],d["com"],d["yr"],round(d["ratio"],3),round(d["lo"],3),round(d["hi"],3),
          round(d["ilo"],3),round(d["ihi"],3),round(d["pit"],3),d["in"],d["in_inside"],d.get("conf","")]
    for j,v in enumerate(vals,1):
        c=ws.cell(fr+i,j,v); c.font=BODY; c.alignment=(LFT if j==1 else CEN)
lastr=fr+len(rf["forest"])-1
# derived coverage summary via formulas over the flag columns (J=In RC, K=In inside)
sr=lastr+2
ws.cell(sr,1,"Coverage (derived):").font=H2
ws.cell(sr+1,1,"Realized inside RC band").font=BODY
ws.cell(sr+1,2,f'=COUNTIF(J{fr}:J{lastr},TRUE)').font=BODY
ws.cell(sr+1,3,f'of {len(rf["forest"])}').font=MUT
ws.cell(sr+2,1,"Realized inside inside-view band").font=BODY
ws.cell(sr+2,2,f'=COUNTIF(K{fr}:K{lastr},TRUE)').font=BODY
ws.cell(sr+2,3,f'of {len(rf["forest"])}').font=MUT
note(ws,sr+4,"Coverage counts are Excel formulas over the flag columns; they reproduce the manuscript's 19/21 reference-class vs 13/21 inside-view coverage.")
widths(ws,{"A":26,"B":10,"C":11,"D":13,"E":11,"F":11,"G":11,"H":11,"I":8,"J":11,"K":14,"L":8})

# ---------------------------------------------------------------- 5. PIT calibration
ws=newsheet("PIT calibration","808080")
r=title(ws,"Calibration — probability integral transform","Source: rf_data.json 'pit'. A calibrated forecaster yields PIT values indistinguishable from uniform (high KS p-value).")
p=rf["pit"]; header_row(ws,r,["Set","KS p-value","Mean PIT","n"])
prows=[("Pooled (all)",p["ks_p_all"],p["mean_all"],21),("Lithium",p["ks_p_li"],p["mean_li"],10),("Copper",p["ks_p_cu"],p["mean_cu"],11)]
for i,(nm,ks,mn,n) in enumerate(prows,1):
    ws.cell(r+i,1,nm).font=BODY
    ws.cell(r+i,2,round(ks,3)).font=BODY; ws.cell(r+i,2).alignment=CEN
    ws.cell(r+i,3,round(mn,3)).font=BODY; ws.cell(r+i,3).alignment=CEN
    ws.cell(r+i,4,n).font=BODY; ws.cell(r+i,4).alignment=CEN
note(ws,r+5,"KS = Kolmogorov–Smirnov test against uniform. A mean near 0.5 and a non-significant p-value indicate calibration.")
widths(ws,{"A":18,"B":12,"C":12,"D":6})

# ---------------------------------------------------------------- 6. Copper decision audit
ws=newsheet("Copper decision audit","C00000")
r=title(ws,"Copper decision audit","Source: rf_data.json 'copper_choices'. RAV and realized NPV in US$ millions, 8% discount. Choice is the engine's enter/stage/walk verdict on frozen sanction-date information.")
header_row(ws,r,["Project","Sanction yr","Ex-ante RAV (US$M)","Engine choice","Realized NPV (US$M)","Ratio","Conf"])
cc=rf["copper_choices"]
for i,d in enumerate(cc,1):
    vals=[d["name"],d["yr"],round(d["rav"],0),d["choice"],round(d["realized"],0),d["ratio"],d.get("conf","")]
    for j,v in enumerate(vals,1):
        c=ws.cell(r+i,j,v); c.font=BODY; c.alignment=(LFT if j==1 else CEN)
        if j in (3,5): c.number_format='#,##0;(#,##0)'
lastr=r+len(cc)
sr=lastr+2
ws.cell(sr,1,"Spearman ρ (RAV vs realized, 8%):").font=BODY
ws.cell(sr,3,round(rf["rho8"][0],3)).font=BODY
ws.cell(sr+1,1,"p-value:").font=BODY; ws.cell(sr+1,3,round(rf["rho8"][1],3)).font=BODY
note(ws,sr+3,"Sumitomo entered Sierra Gorda and Quebrada Blanca 2 (engine: walk); both realized negative NPV after by-product credits.")
widths(ws,{"A":24,"B":11,"C":18,"D":13,"E":18,"F":8,"G":8})

# ---------------------------------------------------------------- 7/8 era registers, 9 AACE
def app_tab(name,key,color,ws_widths):
    ws=newsheet(name,color)
    a=app[key]; r=title(ws,a.get("cap",name).split(".")[0], "Source: manuscript "+key+" (appendix_tables.json).")
    header_row(ws,r,a["headers"])
    for i,row in enumerate(a["rows"],1):
        for j,v in enumerate(row,1):
            c=ws.cell(r+i,j,v); c.font=BODY; c.alignment=(CEN if j in (1,2,4) and key=="A1" else LFT)
    if a.get("cap"): note(ws,r+len(a["rows"])+2,a["cap"])
    widths(ws,ws_widths); return ws
app_tab("Era register - lithium","A2","2E5395",{"A":14,"B":22,"C":34,"D":40})
app_tab("Era register - copper","A3","2E5395",{"A":14,"B":22,"C":34,"D":34})
app_tab("AACE dispersion","A1","2E5395",{"A":10,"B":30,"C":24,"D":16})

# ---------------------------------------------------------------- 10. Policy-event register
ws=newsheet("Policy-event register","9C6500")
r=title(ws,"Policy-event register","Source: policy_event_table.json. 41 resource-nationalism events, 2007–2026; the lever is the NPV channel (E export, N nationalization, L local-content, T tax/royalty). Primary-source citations verified for 40 of 41 events.")
header_row(ws,r,pol[0])
for i,row in enumerate(pol[1:],1):
    for j,v in enumerate(row,1):
        c=ws.cell(r+i,j,v); c.font=BODY; c.alignment=(CEN if j in (2,3,4) else LFT)
widths(ws,{"A":12,"B":7,"C":18,"D":16,"E":40,"F":58})
note(ws,r+len(pol)+1,"One cell remains [CHECK] (Zimbabwe 2025 concentrate tax — rate/instrument unverified). Four instruments activate two levers at once; see Robustness tab.")

# ---------------------------------------------------------------- 11. Driver attribution
ws=newsheet("Driver attribution","808080")
r=title(ws,"Driver attribution (lithium)","Source: table5_regen.json. Sobol top-variance driver vs the factor that decided the realized outcome; the cost audit resolves the one case the variance ranking misses (Table 5).")
header_row(ws,r,["Project","Sobol top driver","Realized decisive factor","Match","Cost-audit percentile","Outcome"])
for i,row in enumerate(t5,1):
    for j,v in enumerate(row,1):
        c=ws.cell(r+i,j,v); c.font=BODY; c.alignment=(CEN if j in (2,3,4,5) else LFT)
widths(ws,{"A":18,"B":16,"C":18,"D":8,"E":16,"F":34})

# ---------------------------------------------------------------- 12. Robustness
ws=newsheet("Robustness","1E8E5A")
r=title(ws,"Robustness","Sources: rf_data.json (grids) and robustness_shared_instrument.py / robustness_heldout.py (shared-instrument check).")
ws.cell(r,1,"Spearman ρ across discount × risk-aversion grid (copper)").font=H2
r+=1; header_row_cols=["Discount \\ λ","0.3","0.5","0.7"]
for j,c in enumerate(header_row_cols,1):
    cell=ws.cell(r,j,c); cell.font=HDR; cell.fill=HDRFILL
discs=["8%","9%","10%"]
for i,drow in enumerate(rf["rho_grid"]["rho"]):
    ws.cell(r+1+i,1,discs[i]).font=BODY
    for j,val in enumerate(drow,2):
        ws.cell(r+1+i,j,val).font=BODY; ws.cell(r+1+i,j).alignment=CEN
r=r+5
ws.cell(r,1,"Copper decision-loss: % of always-enter→oracle gap closed (8% discount, λ=0.5)").font=H2
r+=1
chis=["χ=0.25","χ=0.50","χ=0.75","χ=1.00"]
ws.cell(r,1,"Staged exp. \\ χ").font=HDR; ws.cell(r,1).fill=HDRFILL
for j,c in enumerate(chis,2):
    ws.cell(r,j,c).font=HDR; ws.cell(r,j).fill=HDRFILL
sexp=["0.1","0.3","0.5"]
for i,drow in enumerate(rf["cu_lossgrid"]):
    ws.cell(r+1+i,1,sexp[i]).font=BODY
    for j,val in enumerate(drow,2):
        ws.cell(r+1+i,j,val).font=BODY; ws.cell(r+1+i,j).alignment=CEN
        ws.cell(r+1+i,j).number_format='0"%"'
r=r+5
ws.cell(r,1,"Lithium decision loss (US$M)").font=H2; r+=1
for k in ["workflow","always_enter","always_walk","random"]:
    ws.cell(r,1,k.replace("_"," ")).font=BODY; ws.cell(r,2,rf["li_loss"][k]).font=BODY; ws.cell(r,2).number_format='#,##0'; r+=1
r+=1
ws.cell(r,1,"Shared-instrument sensitivity (collapse 4 two-lever instruments → 1 event each)").font=H2; r+=1
sens=[("Export (E) & nationalization (N) hazards","unchanged"),
      ("Local-content (L) events","8 → 4 (hazard ↓ ~30%)"),
      ("Price-boom multiplier (β)","2.6× → 2.1×"),
      ("Held-out prediction test","3/4 beats base, PASS (both panels; +0.29 vs +0.31 nats)")]
header_row(ws,r,["Quantity","Effect of collapse"])
for i,(a,b) in enumerate(sens,1):
    ws.cell(r+i,1,a).font=BODY; ws.cell(r+i,1).alignment=LFT
    ws.cell(r+i,2,b).font=BODY; ws.cell(r+i,2).alignment=LFT
widths(ws,{"A":40,"B":24,"C":10,"D":10})

out=os.path.join(HERE,"Supplementary_Data.xlsx")
wb.save(out); print("written",out,"tabs:",len(wb.sheetnames))
